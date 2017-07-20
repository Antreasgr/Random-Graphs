import os
import statistics
import sys
import csv
import collections
import yaml
# from yaml import CLoader
from yaml import Loader, Dumper
# from ruamel import yaml
from clique_tree import *
import json
# import xlsxwriter
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.colors import Color
import math


def parse_data(path, output=True):
    all_data = []
    report_path = path + "/Reports"
    files = [file for file in os.listdir(path) if os.path.isfile(os.path.join(path, file))]

    os.makedirs(os.path.dirname(report_path + "/"), exist_ok=True)

    for file in files:
        filename = os.path.splitext(file)[0]
        out_path = os.path.join(report_path, filename)
        if not output or not os.path.exists(out_path + ".csv"):
            print(filename)
            with open(os.path.join(path, file), 'r') as stream:
                yml_data = yaml.load(stream)
                acc = process_data(yml_data)
                all_data.append(acc)
                # print to stdout
                # print_data(yml_data, acc)

                # print csv and md format
                if output:
                    for frmt in ["csv", "md"]:
                        with open(out_path + "." + frmt, 'w') as out_file:
                            print_data(yml_data, acc, out_file, frmt)

    return all_data


def process_data(run_stats):
    accumulative = {"Times": {}, "Stats": {}, "Output": {}}
    for run in run_stats["Run"]:
        for section in accumulative:
            for t in run[section]:
                if not t in accumulative[section]:
                    accumulative[section][t] = []
                if t != "clique_trees":
                    accumulative[section][t].append(run[section][t])
                else:
                    for tree_index, tree in enumerate(run[section][t]):
                        while len(accumulative[section][t]) < tree_index + 1:
                            accumulative[section][t].append({})
                        for treesection in tree:
                            # if treesection == "max_clique_edge_distribution":
                            #     tree[treesection] = (tree["max_size"] * (tree["max_size"] - 1) / 2) / run["Output"]["edges"]

                            if not treesection in accumulative[section][t][tree_index]:
                                accumulative[section][t][tree_index][treesection] = []

                            accumulative[section][t][tree_index][treesection].append(tree[treesection])

    accumulative["Parameters"] = run_stats["Run"][0]["Parameters"]

    return accumulative


def print_data(run_stats, accumulative, report_file=sys.stdout, frmt="csv"):
    if frmt == "md":
        print_header_md(run_stats, report_file)
    else:
        print_header_csv(run_stats, report_file)

    print("## Time - Output Parameters", file=report_file)
    print_mean_std_vertical(((key, value) for section in accumulative for key, value in accumulative[section].items()
                             if key != "clique_trees"), frmt, report_file)

    if "clique_trees" in accumulative["Output"]:
        print("\n" + "-".center(3, "-"), file=report_file)
        print("## Clique trees", file=report_file)
        for tree in accumulative["Output"]["clique_trees"]:
            print('\n', file=report_file)
            print_mean_std_vertical(((o, tree[o]) for o in tree), frmt, report_file)


def print_header_md(run_stats, report_file=sys.stdout):
    print("# Report\n".center(3), file=report_file)
    line1 = "|".join(run_stats["Run"][0]["Parameters"])
    line1 += "|Trials"
    line2 = "|".join(("-" for _ in run_stats["Run"][0]["Parameters"])) + "|-"
    line3 = "|".join([str(run_stats["Run"][0]["Parameters"][param]) for param in run_stats["Run"][0]["Parameters"]])
    line3 += "|" + str(len(run_stats["Run"]))

    print("> |" + line1 + "|", file=report_file)
    print("> |" + line2 + "|", file=report_file)
    print("> |" + line3 + "|", file=report_file)
    print("\n" + "-".center(3, "-"), file=report_file)


def print_header_csv(run_stats, report_file=sys.stdout, delimiter=";"):
    print("Report\n", file=report_file)
    delimiter += " "
    print(delimiter.join(run_stats["Run"][0]["Parameters"]) + delimiter + "Trials", file=report_file)
    print(
        delimiter.join([str(run_stats["Run"][0]["Parameters"][param])
                        for param in run_stats["Run"][0]["Parameters"]]) + delimiter + str(len(run_stats["Run"])),
        file=report_file)
    print("\n" + "-".center(3, "-"), file=report_file)


def print_mean_std_vertical(list_values, frmt="csv", report_file=sys.stdout, delimiter=";"):
    columns = ['', 'mean', 'std']
    prefix = ''
    if frmt == "md":
        delimiter = "|"
        prefix = delimiter
    else:
        delimiter += " "

    print(prefix + delimiter.join(columns) + prefix, file=report_file)
    if frmt == "md":
        print(prefix + delimiter.join(("-" for _ in columns)) + prefix, file=report_file)

    for key, value in list_values:
        mean = "{:>22.15f}".format(statistics.mean(value))
        stddev = "{:>22.15f}".format(statistics.stdev(value))
        p_key = "**" + key + "**" if frmt == "md" else key
        print(prefix + delimiter.join([p_key, mean, stddev]) + prefix, file=report_file)


def print_mean_std(list_values, frmt="csv", report_file=sys.stdout, delimiter=";"):
    lines = ['', '', 'mean', 'std']
    if frmt == "md":
        delimiter = "|"
        lines = [delimiter + l for l in lines]
        lines[1] += "-"
    else:
        delimiter += " "

    for key, value in list_values:
        lines[0] += delimiter + str(key)
        lines[1] += delimiter + "-"
        lines[2] += delimiter + str(statistics.mean(value))
        lines[3] += delimiter + str(statistics.stdev(value))

    if frmt == "md":
        lines = [l + delimiter for l in lines]

    for i, line in enumerate(lines):
        if i != 1 or frmt == "md":
            print(line, file=report_file)


def sort_data_fn(a):
    par = a["Parameters"]
    k = par["k"] if "k" in par else par["edge_density"]
    return par["Algorithm"], par["n"], k


def read_all_data_yml(name, all_data_filename):
    with open(all_data_filename, 'r') as stream:
        all_data = yaml.load(stream, Loader=Loader)

    mva_data = [d for d in all_data if d["Parameters"]["Algorithm"] == name]
    shet_data = []  # [d for d in all_data if d["Parameters"]["Algorithm"] == "SHET"]
    del all_data

    return mva_data, shet_data


def generate_accumulative_report(name, data=None):
    data.sort(key=sort_data_fn)

    lines = []
    for i, datum in enumerate(data):
        ct = datum["Output"]["clique_trees"][0]
        if i == 0:
            header_lines, columns_stats, columns_times, columns_ct = accumulative_header(datum, ct)
            lines.extend(header_lines)

        other_lines = accumulative_line(i, datum, ct, columns_stats, columns_times, columns_ct)
        lines.append(other_lines)

    return lines


def index_of(order_list, value):
    try:
        return order_list.index(value)
    except ValueError:
        return -1


def accumulative_header(datum, ct):
    columns_stats = [o for o in datum["Stats"]]
    columns_times = [o for o in datum["Times"]]

    orderby = TreeStatistics.__slots__
    columns_ct = [o for o in ct if not o.startswith("distribution")]
    columns_ct.sort(key=lambda h: index_of(orderby, h))

    parameters = [pp for pp in datum["Parameters"] if (pp != "Algorithm" and pp != "seed" and pp != "version"  and pp != "k")]

    l_1 = [datum["Parameters"]["Algorithm"]]

    l_11 = ["" for pp in parameters] + ["Stats"] + ["" for _ in range(2 * len(columns_stats) - 1)]
    l_11 += ["Times"] + ["" for _ in range(2 * len(columns_times) - 1)]
    l_11 += ["Clique Tree"] + ["" for _ in range(2 * len(columns_ct) - 1)]

    l_2 = ["" for pp in parameters]
    l_2 += [header if not twice else "" for header in columns_stats for twice in range(2)]
    l_2 += [header if not twice else "" for header in columns_times for twice in range(2)]
    l_2 += [header if not twice else "" for header in columns_ct for twice in range(2)]

    l_3 = [pp for pp in parameters]
    l_3 += [twice for header in columns_stats for twice in ["mean", "std"]]
    l_3 += [twice for header in columns_times for twice in ["mean", "std"]]
    l_3 += [twice for header in columns_ct for twice in ["mean", "std"]]

    return [l_1, l_11, l_2, l_3], columns_stats, columns_times, columns_ct


def accumulative_line(i, datum, ct, columns_stats, columns_times, columns_ct):
    par = datum["Parameters"]
    stats = datum["Stats"]
    times = datum["Times"]

    l = [str(par[pp]) for pp in par if (pp != "Algorithm" and pp != "seed" and pp != "version" and pp != "k")]

    l.extend([fn(stats[k]) if k in stats else "" for k in columns_stats for fn in [statistics.mean, statistics.stdev]])

    l.extend([fn(times[k]) if k in times else "" for k in columns_times for fn in [statistics.mean, statistics.stdev]])

    all_values = (ct[k] for k in columns_ct if len(ct[k]) and not isinstance(ct[k][0], dict))
    sanitize_values = []
    for array_of_values in all_values:
        col = [vv for vv in array_of_values if not isinstance(vv, str) and not math.isinf(vv) and not math.isnan(vv)]
        if col:
            sanitize_values.append(col)

    if sanitize_values:
        l.extend([fn(v) if len(v) > 1 else v[0] for v in sanitize_values for fn in [statistics.mean, statistics.stdev]])
    else:
        l.extend([0, 0])

    return l


def localize_floats(row):
    return [str(el).replace('.', ',') if isinstance(el, float) else el for el in row]


def run_reports_data(name, data=[]):
    if data:
        with open(os.path.join("Results", "all_data_" + name + ".yml"), 'w') as stream:
            yaml.dump(data, stream)

    # mva_data, shet_data = read_all_data_yml(name, os.path.join("Results", "all_data_" + name + ".yml"))
    all_lines = generate_accumulative_report(name, data)
    print(all_lines)
    if all_lines:
        write_excel(all_lines, os.path.join("Results", name + ".xlsx"), ['Title', 'Headline 1', 'Headline 2', 'Headline 3'])


def run_reports(name):
    # mva_data = parse_data("Results/" + name, False)
    # shet_data = []  # parse_data("Results/" + name, False)
    with open(os.path.join("Results", "all_data_" + name + ".yml"), 'r') as stream:
        mva_data = yaml.load(stream)

    print("Done...")

    run_reports_data(name, mva_data)


def excel_reports(algorithms):
    # read ALL data
    all_data = {}
    # for name in algorithms:
    #     all_data_filename = os.path.join("Results", "all_data_" + name + ".yml")
    #     with open(all_data_filename, 'r') as stream:
    #         all_data[name] = yaml.load(stream, Loader=Loader)

    #     all_data[name].sort(key=sort_data_fn)

    # print(all_data.keys())

    # with open(os.path.join("Results", "all_data.json"), 'w') as stream:
    #         json.dump(all_data, stream)

    with open(os.path.join("Results", "all_data.json"), 'r') as stream:
        all_data = json.load(stream)

    for name in algorithms:
        all_data[name].sort(key=sort_data_fn)
        for d in all_data[name]:
            if "edge_density" in d["Output"]:
                d["Stats"]["edge_density"] = d["Output"]["edge_density"]
            elif "actual_edge_density" in d["Stats"]:
                d["Stats"]["edge_density"] = d["Stats"]["actual_edge_density"]
            elif "edge_density" not in d["Stats"]:
                d["Stats"]["edge_density"] = [d["Parameters"]["edge_density"]]

    # get headers from the first algorithm
    stats_headers = [h for h in all_data[algorithms[0]][0]["Stats"]]
    stats_headers.sort()
    ct_headers = [h for h in all_data[algorithms[0]][0]["Output"]["clique_trees"][-1] if not h.startswith("distribution")]

    orderby = TreeStatistics.__slots__
    ct_headers.sort(key=lambda h: index_of(orderby, h))

    rows = []
    rows.append(["n"] + [h for h in stats_headers for name in algorithms] + [h for h in ct_headers for name in algorithms])
    rows.append([""] + [name for h in stats_headers for name in algorithms] + [name for h in ct_headers for name in algorithms])

    indexes = [(name, 0) for name in algorithms]
    while [j < len(all_data[name]) for name, j in indexes].count(True) == len(indexes):
        rown = all_data[indexes[0][0]][indexes[0][1]]["Parameters"]["n"]

        # check at least n is same for all algorithms
        valid_for_algorithm = []
        for name, i in indexes:
            if all_data[name][i]["Parameters"]["n"] != rown:
                # raise Exception("{0} != {1}, missing row".format(all_data[name][i]["Parameters"]["n"], rown))
                print("{0} in {2} != {1} in {3}, missing row".format(all_data[name][i]["Parameters"]["n"], rown, name, indexes[0][0]))
            else:
                valid_for_algorithm.append(name)

        stat_values = []
        for stat in stats_headers:
            for name, i in indexes:
                if name in valid_for_algorithm:
                    if stat in all_data[name][i]["Stats"]:
                        stat_values.append(statistics.mean(all_data[name][i]["Stats"][stat]))
                        continue
                stat_values.append("")

        ct_values = []
        for h in ct_headers:
            for name, i in indexes:
                if name in valid_for_algorithm:
                    # sanitize for -inf, inf
                    values_10 = [vv for vv in all_data[name][i]["Output"]["clique_trees"][-1][h] if not isinstance(vv, str)]
                    if values_10:
                        col_value = statistics.mean(values_10)
                    else:
                        # all values are inf
                        col_value = 'inf'
                    ct_values.append(col_value)
                else:
                    ct_values.append("")

        rows.append([rown] + stat_values + ct_values)

        indexes = [(name, i + 1) if name in valid_for_algorithm else (name, i) for name, i in indexes]

    print(rows)
    write_excel(rows, os.path.join("Results", "comparison.xlsx"))


def write_excel(rows, filename, header_rows=['Headline 2', 'Headline 3']):
    workbook = Workbook()
    worksheet = workbook["Sheet"]

    col_widths = [max(8, len(v) + 4) for v in rows[0]]
    for i, _ in enumerate(rows[0]):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = col_widths[i]

    for row_index, row in enumerate(rows):
        if row_index < len(header_rows):
            cells = [WriteOnlyCell(worksheet, value=v) for v in row]
            for cell in cells:
                cell.style = header_rows[row_index]
            worksheet.append(cells)
        else:
            worksheet.append(row)

    workbook.save(filename)


NAME = "SHET_PRUNED"
if __name__ == '__main__':
    run_reports(NAME)
    # excel_reports(["SHET", "MVA2", "INCR_k_1e-5", "INCR_k_1_rev_2"])