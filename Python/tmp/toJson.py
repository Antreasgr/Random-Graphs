﻿# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import pprint
import json
import sys


def check_add_node(parent, name, description, size = 1):
    field = [c for c in parent["children"] if c and c["name"] == name]
    if len(field) == 0:
        field = {"name": name, "description": description, "size": size, "children": []}
        parent["children"].append(field)
        return field
    return field[0]


wb = load_workbook(filename='C:/Users/Antreas/Downloads/sxoles.xlsx')
result_dict = { "name": "flare", "description": "flare", "children": [] }

for sheet_name in wb.sheetnames:
    parent = check_add_node(result_dict, sheet_name, sheet_name)
    for i, row in enumerate(wb[sheet_name].iter_rows()):
        if i < 2:
            continue

        ep_pedia = row[1].value.split(",")
        tmima = row[2].value.split(",")[0] if row[2].value else ""
        idr = row[3].value
        polh = row[4].value
        eidiko_mathima = row[5].value != None
        vasi = row[7].value

        if tmima == None or polh == None:
            continue

        for pedio in ep_pedia:
            pedio_field = check_add_node(parent, pedio, pedio)
            idr_field = check_add_node(pedio_field, idr, idr)
            tmima_field = check_add_node(idr_field, tmima, tmima)
            polh_field = check_add_node(tmima_field, polh, tmima + " " + polh + ", " + str(vasi), 1)



# for pedio in result_dict["children"][0]["children"]:
#     for idr in pedio["children"]:
#         for tmima in idr["children"]:
#             if len(tmima["children"]) == 1:
#                 tmima["description"] += ", " + tmima["children"][0]["name"]
#             tmima["children"] = []


# pprint.pprint(result_dict)
print(json.dumps(result_dict, indent=4, ensure_ascii=False, sort_keys=True))
with open("sxoles.json", "w", encoding='utf8') as stream:
    json.dump(result_dict, stream, ensure_ascii=False) 
# json.dump(result_dict, sys.stdout)
